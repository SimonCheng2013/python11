# -*- conding: utf-8 -*-
# @Time      :2019/7/15 15:17
# Author     :chengjin
# @Email     :amazing2013@163.com
# @File      :class_excel.py
'''地址 测试数据 断言期望结果 除了这几个不同 其他高度相似'''
'''参数化 url data status'''

from openpyxl import load_workbook
#打开表单
wb = load_workbook("xh.xlsx")
#2.定位表单
sheet = wb["python"]

#3.定位单元格 行列值
res = sheet.cell(1,1).value
res2 = sheet.cell(2,1).value


# print("最大行：{}".format(sheet.max_row))
# print("最大列：{}".format(sheet.max_column))

# print("url:{}".format(sheet.cell(1,1).value))
# print("data:{}".format(sheet.cell(1,2).value))
# print("url:{}".format(sheet.cell(1,3).value))
# print("url:{}".format(sheet.cell(1,4).value))

'''eval 把数据类转换成 原本数据类型'''
# s = 'True'
s ='{"age":18}'
# print(s,type(s))

print(eval(s),type(eval(s)))

