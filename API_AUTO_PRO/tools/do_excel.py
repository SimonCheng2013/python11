# -*- conding: utf-8 -*-
# @Time      :2019/7/26 16:36
# Author     :chengjin
# @Email     :amazing2013@163.com
# @File      :do_excel.py
from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools import project_path
from tools.get_data import GetData
class DoExcel:
    @classmethod
    def get_data(cls,file_name):
        wb = load_workbook(file_name)  # linux命令
        mode =eval(ReadConfig.get_config(project_path.case_config_path,'MODE','mode'))
        tel = getattr(GetData,"NoRegTel")#从GetData拿到的手机号
        test_data = []
        for key in mode:#遍历这个存在配置文件里面的字典
            sheet = wb[key]#表单名
            if mode[key]=='all':
                for i in range(2,sheet.max_row+1):
                    row_data ={}#字典
                    row_data["case_id"]=sheet.cell(i,1).value
                    row_data["url"]=sheet.cell(i,2).value
                    row_data["data"]=sheet.cell(i,3).value
                    # if sheet.cell(i,3).value.find("${tel_1}")!=-1:
                    #     row_data["data"] = sheet.cell(i, 3).value.replace("${tel_1}",tel)
                    # elif sheet.cell(i,3).value.find("${tel}")!=-1:
                    #     row_data["data"] = sheet.cell(i, 3).value.replace("${tel_1}",tel)
                    # else:
                    #     row_data["data"]=sheet.cell(i,3).value
                    row_data["title"]=sheet.cell(i,4).value
                    row_data["http_method"]=sheet.cell(i,5).value
                    row_data["excepted"]=sheet.cell(i,6).value
                    row_data["sheet_name"]=key
                    test_data.append(row_data)
                    # cls.updata_tel(tel+2,file_name,"init")#更新手机号
                    # cls.updata_tel(tel,file_name,"init")#更新手机号
            else:
                for case_id in mode[key]:
                    row_data = {}  # 字典
                    row_data["case_id"] = sheet.cell(case_id+1, 1).value#行号
                    row_data["url"] = sheet.cell(case_id+1, 2).value
                    # row_data["data"] = sheet.cell(case_id+1, 3).value
                    if sheet.cell(i,3).value.find("${tel_1}")!=-1:
                        row_data["data"] = sheet.cell(i, 3).value.replace("${tel_1}",tel)
                    elif sheet.cell(i,3).value.find("${tel}")!=-1:
                        row_data["data"] = sheet.cell(i, 3).value.replace("${tel_1}",tel)
                    else:
                        row_data["data"]=sheet.cell(i,3).value
                    row_data["title"] = sheet.cell(case_id+1, 4).value
                    row_data["http_method"] = sheet.cell(case_id+1, 5).value
                    row_data["excepted"] = sheet.cell(case_id+1, 6).value
                    test_data.append(row_data)
                    cls.updata_tel(tel+2,file_name,"init")#更新手机号
        return test_data
    @staticmethod
    def write_back(file_name,sheet_name,i,result,TestResult):#专门写回数据
        wb=load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,7).value=result
        sheet.cell(i,8).value=TestResult
        wb.save(file_name)#保存结果
    @classmethod
    def updata_tel(cls,tel,filename,sheet_name):
        '''更新Excel里面的数据'''
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(filename)
if __name__ == '__main__':
    test_data=DoExcel().get_data(project_path.test_case_path)
    print(test_data)