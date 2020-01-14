# -*- conding: utf-8 -*-
# @Time      :2019/8/13 16:33
# Author     :chengjin
# @Email     :amazing2013@163.com
# @File      :do_excel.py
from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools import project_path
from tools.get_data import GetData
class DoExcel:
    @classmethod
    def get_data(cls,file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(project_path.case_config_path,"MODE","mode"))
        tel =getattr(GetData,"NoRegTel")
        test_data= []
        for key in mode:#遍历这个存在配置文件里面的字典
            sheet = wb[key]#表单名
            if mode[key]=="all":
                for i in range(2,sheet.max_row+1):
                    row_data = {}#字典
                    row_data["case_id"] = sheet.cell(i, 1).value
                    row_data["model"] = sheet.cell(i, 2).value
                    row_data["title"] = sheet.cell(i, 3).value
                    row_data["url"] = sheet.cell(i, 4).value
                    row_data["params"] = sheet.cell(i, 5).value
                    row_data["header"] = sheet.cell(i, 6).value
                    row_data["http_method"] = sheet.cell(i, 7).value
                    row_data["data"] = sheet.cell(i, 8).value
                    row_data["excepted_code"] = sheet.cell(i, 9).value
                    row_data["sheet_name"] = key
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}  # 字典
                    row_data["case_id"] = sheet.cell(case_id+1, 1).value#行号
                    row_data["model"] = sheet.cell(case_id+1, 2).value
                    row_data["title"] = sheet.cell(case_id+1, 3).value
                    row_data["url"] = sheet.cell(case_id+1, 4).value
                    row_data["params"] = sheet.cell(case_id+1, 5).value
                    row_data["header"] = sheet.cell(case_id+1, 6).value
                    row_data["http_method"] = sheet.cell(case_id+1, 7).value
                    row_data["data"] = sheet.cell(case_id+1, 8).value
                    row_data["excepted_code"] = sheet.cell(case_id+1, 9).value


                    test_data.append(row_data)
                    # cls.updata_tel(tel+2,file_name,"init")#更新手机号
        return test_data
    @staticmethod
    def write_back(file_name,sheet_name,i,result,TestResult):#写回数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 10).value = result
        sheet.cell(i, 11).value = TestResult
        wb.save(file_name)  # 保存结果
if __name__ == '__main__':
    test_data=DoExcel().get_data(project_path.test_case_path)
    print(test_data)
