# -*- conding: utf-8 -*-
# @Time      :2019/7/16 17:57
# Author     :chengjin
# @Email     :amazing2013@163.com
# @File      :do_excel.py
from openpyxl import load_workbook
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]
        self.max_row = self.sheet_obj.max_row
        #获取一个表单对象
    def get_data(self,i,j):
        '''根据传入的坐标来获取值'''
        return self.sheet_obj.cell(i,j).value

if __name__ == '__main__':
    res = DoExcel("xh.xlsx","python").get_data(1,1)
    print(res)