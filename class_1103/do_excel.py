# -*- conding: utf-8 -*-
# @Time      :2019/7/16 17:57
# Author     :chengjin
# @Email     :amazing2013@163.com
# @File      :do_excel.py
from openpyxl import load_workbook
from class_1103.read_config import ReadConfig
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_data(self):
        '''button:控制是否执行所有用例，默认为all 为all就执行所有用例
        如果不为all就不全部执行
        mode的值 只能输入 all 列表 这两种类型的参数
        '''
        mode = ReadConfig().read_config('case.config',"MODE","mode")
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data=[]
        for i in range(1,sheet.max_row+1):
            sub_data={}
            sub_data["case_id"] = sheet.cell(i, 1).value
            sub_data["module"] = sheet.cell(i, 2).value
            sub_data["title"] = sheet.cell(i, 3).value
            sub_data["method"] = sheet.cell(i, 4).value
            sub_data["url"]=sheet.cell(i,5).value
            sub_data["data"]=sheet.cell(i,6).value
            sub_data["status"]=sheet.cell(i,7).value
            test_data.append(sub_data)#储存所有数据

        #根据mode值进行判断
        if mode=="all":
            final_data=test_data
        else:#[1,2,3,4]
            final_data=[]
            for item in test_data:
                if item["case_id"] in eval(mode):
                    final_data.append(item)
        return final_data
if __name__ == '__main__':
    print(DoExcel("xh.xlsx","python").get_data())