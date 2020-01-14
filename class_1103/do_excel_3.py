# -*- conding: utf-8 -*-
# @Time      :2019/7/16 17:57
# Author     :chengjin
# @Email     :amazing2013@163.com
# @File      :do_excel.py
from openpyxl import load_workbook
#方法三：仅供参考 日后可以拿来优化
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_header(self):
        '''获取第一行的标题行'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header
    def get_data(self):
        '''根据嵌套循环读取数据'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        header = self.get_header()#拿到header 是一个索引从0开始的
        print("header",header)
        test_data=[]

        for i in range(2,sheet.max_row+1):# 1 2 3 4
            sub_data={}
            for j in range(1,sheet.max_column+1):# 1 2 3 4 5 6 7
                sub_data[header[j-1]]=sheet.cell(i,j).value
            test_data.append(sub_data)
        return test_data
if __name__ == '__main__':
    print(DoExcel("xh.xlsx","python").get_data())